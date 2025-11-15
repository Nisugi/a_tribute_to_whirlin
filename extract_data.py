"""
Extract data from Excel spreadsheet to JSON format
"""
import openpyxl
import json
import sys
from pathlib import Path

def extract_sheet_data(sheet):
    """Extract all data from a worksheet including merged cells"""
    data = {
        'name': sheet.title,
        'rows': [],
        'merged_cells': [],
        'dimensions': {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column
        }
    }

    # Extract merged cell ranges
    for merged_range in sheet.merged_cells.ranges:
        data['merged_cells'].append(str(merged_range))

    # Extract cell data
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            cell_info = {
                'value': cell.value,
                'coordinate': cell.coordinate,
            }

            # Add formatting info if available
            if cell.font:
                cell_info['font'] = {
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': str(cell.font.color) if cell.font.color else None
                }

            if cell.fill:
                cell_info['fill'] = str(cell.fill.fgColor) if cell.fill.fgColor else None

            row_data.append(cell_info)

        data['rows'].append(row_data)

    return data

def main():
    # Load the workbook
    excel_file = Path(r'C:\Gemstone\Whirlin_Trainer_2025.1.VBA.xlsm')
    output_dir = Path(r'C:\gemstone\projects\whirlin_spreadsheet\data')

    print(f"Loading workbook: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # Extract each sheet
    all_sheets_data = {}

    for sheet_name in wb.sheetnames:
        print(f"Extracting sheet: {sheet_name}")
        sheet = wb[sheet_name]
        sheet_data = extract_sheet_data(sheet)

        # Save individual sheet as JSON
        sheet_filename = sheet_name.replace(' ', '_').replace('/', '_') + '.json'
        output_file = output_dir / sheet_filename

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, indent=2, default=str)

        print(f"  Saved to: {output_file}")

        # Add to combined data
        all_sheets_data[sheet_name] = sheet_data

    # Save combined data
    combined_file = output_dir / 'all_sheets.json'
    with open(combined_file, 'w', encoding='utf-8') as f:
        json.dump(all_sheets_data, f, indent=2, default=str)

    print(f"\nAll data extracted successfully!")
    print(f"Individual sheets: {len(wb.sheetnames)} files")
    print(f"Combined file: {combined_file}")

    # Create a summary
    summary = {
        'total_sheets': len(wb.sheetnames),
        'sheets': [
            {
                'name': name,
                'rows': all_sheets_data[name]['dimensions']['max_row'],
                'columns': all_sheets_data[name]['dimensions']['max_column']
            }
            for name in wb.sheetnames
        ]
    }

    summary_file = output_dir / 'summary.json'
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2)

    print(f"Summary saved to: {summary_file}")

if __name__ == '__main__':
    main()
